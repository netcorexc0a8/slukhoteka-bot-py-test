from typing import List
from io import BytesIO
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from sqlalchemy.orm import Session

from app.models.booking import Booking, BookingStatus, BookingType
from app.models.user import GlobalUser, PlatformUser


def _service_label(booking: Booking) -> str:
    if booking.service is None:
        return ""
    return booking.service.name


def _booking_type_label(booking: Booking) -> str:
    bt = booking.booking_type
    if bt == BookingType.GROUP or (isinstance(bt, str) and bt == "group"):
        return "Групповое"
    return "Индивидуальное"


class ExcelExportService:
    """Экспорт расписания в Excel."""

    def _merge_same_cells(self, ws, start_row, end_row, column_idx, alignment):
        current_value = None
        merge_start = None
        for row in range(start_row, end_row + 1):
            cell = ws.cell(row=row, column=column_idx)
            cell_value = cell.value
            if cell_value != current_value:
                if merge_start is not None and merge_start != row - 1:
                    ws.merge_cells(
                        start_row=merge_start, start_column=column_idx,
                        end_row=row - 1, end_column=column_idx,
                    )
                merge_start = row
                current_value = cell_value
        if merge_start is not None and merge_start != end_row:
            ws.merge_cells(
                start_row=merge_start, start_column=column_idx,
                end_row=end_row, end_column=column_idx,
            )
        for row in range(start_row, end_row + 1):
            ws.cell(row=row, column=column_idx).alignment = alignment

    def _get_platform_by_specialist(self, db: Session, specialist_id: int) -> str:
        platform_user = (
            db.query(PlatformUser).filter(PlatformUser.global_user_id == specialist_id).first()
        )
        if platform_user:
            return "Telegram" if platform_user.platform == "telegram" else "VK"
        return ""

    def _write_sheet(self, ws, bookings: List[Booking], db: Session):
        headers = [
            "Дата записи", "День недели", "Время записи",
            "Клиент", "Специалист", "Со-ведущие",
            "Тип занятия", "Группа",
            "Абонемент", "Занятие",
            "Статус", "Платформа записи", "Дата создания", "Дата обновления",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        days_of_week = [
            "Понедельник", "Вторник", "Среда", "Четверг",
            "Пятница", "Суббота", "Воскресенье",
        ]

        for row, b in enumerate(bookings, 2):
            client_name = b.client.name if b.client else ""
            specialist_name = (
                b.specialist.name or b.specialist.phone
                if b.specialist else ""
            )
            session_label = ""
            if b.session_number and b.subscription:
                session_label = f"{b.session_number}/{b.subscription.total_sessions}"

            co_specialists = list(b.co_specialists or [])
            co_names = [
                (cs.name or cs.phone or "") for cs in co_specialists
                if cs.id != b.specialist_id
            ]
            co_label = ", ".join(n for n in co_names if n)

            group_name = b.group.name if b.group else ""

            ws.cell(row=row, column=1, value=b.start_time.strftime("%d.%m.%Y"))
            ws.cell(row=row, column=2, value=days_of_week[b.start_time.weekday()])
            ws.cell(row=row, column=3, value=b.start_time.strftime("%H:%M"))
            ws.cell(row=row, column=4, value=client_name)
            ws.cell(row=row, column=5, value=specialist_name)
            ws.cell(row=row, column=6, value=co_label)
            ws.cell(row=row, column=7, value=_booking_type_label(b))
            ws.cell(row=row, column=8, value=group_name)
            ws.cell(row=row, column=9, value=_service_label(b))
            ws.cell(row=row, column=10, value=session_label)
            ws.cell(row=row, column=11, value=b.status.value if b.status else "")
            ws.cell(row=row, column=12, value=self._get_platform_by_specialist(db, b.specialist_id))
            ws.cell(row=row, column=13, value=b.created_at.strftime("%d.%m.%Y %H:%M"))
            ws.cell(row=row, column=14, value=b.updated_at.strftime("%d.%m.%Y %H:%M"))

        if bookings:
            end_row = len(bookings) + 1
            self._merge_same_cells(
                ws, 2, end_row, 1, Alignment(horizontal="left", vertical="top")
            )
            date_groups = defaultdict(list)
            for row, b in enumerate(bookings, 2):
                date_groups[b.start_time.strftime("%d.%m.%Y")].append(row)
            for date, rows in date_groups.items():
                if len(rows) > 1:
                    ws.merge_cells(
                        start_row=rows[0], start_column=2,
                        end_row=rows[-1], end_column=2,
                    )
                    for r in rows:
                        ws.cell(r, 2).alignment = Alignment(horizontal="left", vertical="top")

        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

    def export_schedule(
        self,
        bookings: List[Booking],
        current_user: GlobalUser,
        db: Session,
    ) -> bytes:
        bookings = [b for b in bookings if b.deleted_at is None]

        wb = Workbook()
        ws_general = wb.active
        ws_general.title = "Общее расписание"
        self._write_sheet(ws_general, bookings, db)

        self._add_statistics_sheets(wb, bookings, db)

        per_specialist = defaultdict(list)
        for b in bookings:
            if b.specialist:
                name = b.specialist.name or b.specialist.phone
                per_specialist[name].append(b)
        for name, items in per_specialist.items():
            ws_spec = wb.create_sheet(name[:31])
            self._write_sheet(ws_spec, items, db)

        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    def _add_statistics_sheets(self, wb: Workbook, bookings: List[Booking], db: Session):
        from app.crud.user import get_all_users
        users = get_all_users(db)

        daily_by_user = defaultdict(lambda: defaultdict(int))
        weekly_by_user = defaultdict(lambda: defaultdict(int))

        for b in bookings:
            if not b.specialist:
                continue
            date = b.start_time.date()
            user_name = b.specialist.name or b.specialist.phone
            daily_by_user[user_name][str(date)] += 1
            iso = date.isocalendar()
            week = f"{iso[0]}-W{iso[1]:02d}"
            weekly_by_user[user_name][week] += 1

        users_with_records = set(daily_by_user.keys()) | set(weekly_by_user.keys())
        user_names = [
            u.name or u.phone for u in users
            if (u.name or u.phone) in users_with_records
        ]

        ws_combined = wb.create_sheet("Статистика по дням-неделям")

        daily_dates = sorted({d for ud in daily_by_user.values() for d in ud.keys()})
        if daily_dates:
            ws_combined.cell(row=1, column=1, value="Дата").font = Font(bold=True)
            for col, name in enumerate(user_names, 2):
                ws_combined.cell(row=1, column=col, value=name).font = Font(bold=True)
            total_col = len(user_names) + 2
            ws_combined.cell(row=1, column=total_col, value="Всего занятий").font = Font(bold=True)
            for row, date_str in enumerate(daily_dates, 2):
                ws_combined.cell(row=row, column=1, value=date_str)
                daily_total = 0
                for col, name in enumerate(user_names, 2):
                    cnt = daily_by_user[name].get(date_str, 0)
                    ws_combined.cell(row=row, column=col, value=cnt)
                    daily_total += cnt
                ws_combined.cell(row=row, column=total_col, value=daily_total)

        weekly_weeks = sorted({w for uw in weekly_by_user.values() for w in uw.keys()})
        if weekly_weeks:
            start_col = len(user_names) + 5
            ws_combined.cell(row=1, column=start_col, value="Неделя").font = Font(bold=True)
            for col, name in enumerate(user_names, start_col + 1):
                ws_combined.cell(row=1, column=col, value=name).font = Font(bold=True)
            weekly_total_col = start_col + len(user_names) + 1
            ws_combined.cell(row=1, column=weekly_total_col, value="Всего занятий").font = Font(bold=True)
            for row, week in enumerate(weekly_weeks, 2):
                ws_combined.cell(row=row, column=start_col, value=week)
                weekly_total = 0
                for col, name in enumerate(user_names, start_col + 1):
                    cnt = weekly_by_user[name].get(week, 0)
                    ws_combined.cell(row=row, column=col, value=cnt)
                    weekly_total += cnt
                ws_combined.cell(row=row, column=weekly_total_col, value=weekly_total)

        for col in ws_combined.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws_combined.column_dimensions[col[0].column_letter].width = max_length + 2